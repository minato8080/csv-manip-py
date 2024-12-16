import csv
import os
from openpyxl import Workbook
from openpyxl.comments import Comment
from datetime import datetime, timedelta
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


class CsvToExcelConverter(object):
    def __init__(self):
        self.default_excel_file = None
        self.csv_file = None
        self.file_label = None
        self.info_label = None
        self.tree = None
        self.encoding_select = None
        self.header_var = None
        self.max_columns = 0
        self.date_select = None
        self.pay_select = None
        self.comment_select = None
        self.template_var = None

        return super().__init__()

    def convert_csv_to_json(self, reader):
        # CSVの1列目の日付を収集し、日付以外のデータを弾く
        dates = []
        valid_rows = []
        for row in reader:
            try:
                # 日付をdatetimeオブジェクトに変換
                date_obj = datetime.strptime(
                    row[int(self.date_select.get()) - 1], "%Y/%m/%d"
                )
                dates.append(date_obj)
                valid_rows.append(row)
            except ValueError:
                # 日付以外のデータが入っていた場合は無視
                continue

        # 最小値と最大値を取得
        if not dates:
            return json.dumps({}, ensure_ascii=False, indent=4)

        min_date = min(dates)
        max_date = max(dates)

        # 日付範囲を作成
        date_range = [
            min_date + timedelta(days=x) for x in range((max_date - min_date).days + 1)
        ]

        # 日付をキーにしてデータをまとめる
        date_dict = {date.strftime("%Y/%m/%d"): [] for date in date_range}
        for row in valid_rows:
            date_str = datetime.strptime(
                row[int(self.date_select.get()) - 1], "%Y/%m/%d"
            ).strftime("%Y/%m/%d")
            if date_str in date_dict:
                date_dict[date_str].append(row[1:])

        # JSON形式にパース
        return json.dumps(date_dict, ensure_ascii=False, indent=4)

    def csv_to_excel_with_comments(self, excel_file):
        # CSVファイルを読み込む
        encoding = self.encoding_select.get()
        with open(self.csv_file, mode="r", encoding=encoding) as f:
            # CSVリーダーを作成
            reader = csv.reader(f)
            # ヘッダ行の有無
            if self.header_var.get():
                next(reader)

            # CSVデータをフォーマットしてJSON形式に変換
            json_data = json.loads(self.convert_csv_to_json(reader))

        # Excelファイルを作成
        wb = Workbook()
        ws = wb.active

        # データをExcelに書き込み
        try:
            for col_idx, key in enumerate(json_data, start=1):
                cell = ws.cell(
                    row=1,
                    column=col_idx,
                    value=datetime.strptime(key, "%Y/%m/%d").strftime("%m/%d"),
                )
                for row_idx, row in enumerate(json_data[key], start=2):
                    cell = ws.cell(
                        row=row_idx,
                        column=col_idx,
                        value=f"¥{int(row[int(self.pay_select.get()) - 2]):,}",
                    )
                    # コメントを追加
                    comment_text = f"{row[int(self.comment_select.get()) - 2]}"
                    cell.comment = Comment(comment_text, "GeneratedByScript")
        except ValueError as e:
            self.info_label.config(text=f"日付または金額フォーマットが不正です。")
            return

        # 列幅を1.2cmに設定
        for col in ws.columns:
            max_length = 1.2 / 0.18  # 1.2cmをポイントに変換（1ポイント=0.18cm）
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max_length

        # Excelファイルを保存
        wb.save(excel_file)
        messagebox.showinfo("完了", f"Excelファイルを作成しました: {excel_file}")

    def csv_import(self, is_open):
        self.info_label.config(text="")
        if is_open:
            self.csv_file = filedialog.askopenfilename(
                title="CSVファイルを選択", filetypes=[("CSV files", "*.csv")]
            )
        if not self.csv_file:
            return
        # デフォルトで選択したCSVファイルと同じ名前のExcelファイルを設定（フルパスを除去）
        self.default_excel_file = (
            os.path.basename(self.csv_file).rsplit(".", 1)[0] + ".xlsx"
        )
        self.file_label.config(text=f"{self.csv_file}")

        for item in self.tree.get_children():
            self.tree.delete(item)  # 既存の内容をクリア

        encoding = self.encoding_select.get()
        try:
            with open(self.csv_file, mode="r", encoding=encoding) as f:
                reader = csv.reader(f)
                if self.header_var.get():
                    next(reader)
                max_columns = 0
                # Treeviewに挿入
                for row in reader:
                    if max_columns < len(row):
                        max_columns = len(row)
                    self.tree.insert("", "end", values=row)

                # ヘッダーを設定
                columns = [f"#{i}" for i in range(1, max_columns + 1)]
                self.tree["columns"] = columns
                for idx in range(1, max_columns + 1):
                    self.tree.heading(f"#{idx}", text=idx, anchor="w")
                    self.tree.column(f"#{idx}", anchor="w")

                self.max_columns = max_columns
                self.date_select["values"] = [str(i) for i in range(1, max_columns + 1)]
        except UnicodeDecodeError as e:
            self.info_label.config(text=f"文字コードが不正です。")

    def excel_export(self):
        self.info_label.config(text="")
        if not self.csv_file:
            self.info_label.config(text="ファイルを選択していません。")
            return
        excel_file = filedialog.asksaveasfilename(
            title="Excelファイルを保存",
            initialfile=self.default_excel_file,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if not excel_file:
            return
        self.csv_to_excel_with_comments(excel_file)

    def on_template_change(self, _event):
        template = self.template_var.get()
        if template == "三井住友":
            # 三井住友のテンプレートに対する処理
            self.encoding_select.set("shift_jis")
            self.date_select.set("1")
            self.comment_select.set("2")
            self.pay_select.set("3")
            self.csv_import(False)
            pass
        elif template == "amazon":
            # amazonのテンプレートに対する処理
            self.encoding_select.set("utf_8")
            self.date_select.set("1")
            self.comment_select.set("3")
            self.pay_select.set("8")
            self.csv_import(False)
            pass
        else:
            # その他のテンプレートに対する処理
            pass

    def init(self):
        # メインウィンドウを作成
        root = tk.Tk()
        root.title("CSV to Excel Converter")

        # 各種フレームを作成
        frame_button = tk.Frame(root, width=80, height=300)
        frame_button.grid(padx=2, pady=2, sticky="w")

        frame_manip = tk.Frame(root, width=80, height=300)
        frame_manip.grid(padx=2, pady=2, sticky="w")

        frame_info = tk.Frame(root, width=80, height=300)
        frame_info.grid(padx=2, pady=2, sticky="w")

        frame_tree = tk.Frame(root, width=620, height=200)
        frame_tree.grid(padx=2, pady=2, sticky="w")

        # CSVインポートボタンを作成
        import_button = tk.Button(
            frame_button, text="CSV Import", command=lambda: self.csv_import(True)
        )
        import_button.grid(row=0, column=0, padx=5, pady=5, sticky="w")

        # Excelエクスポートボタンを作成
        export_button = tk.Button(
            frame_button, text="Excel Export", command=self.excel_export
        )
        export_button.grid(row=0, column=1, padx=5, pady=5, sticky="w")

        # エンコーディング選択ボックスを作成
        encoding_var = tk.StringVar(value="shift_jis")
        self.encoding_select = ttk.Combobox(
            frame_button,
            textvariable=encoding_var,
            values=["utf_8", "shift_jis"],
            width=8,
        )
        self.encoding_select.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.encoding_select.bind(
            "<<ComboboxSelected>>", lambda event: self.csv_import(False)
        )

        # ヘッダ行の有無を選択するラジオボタンを作成
        self.header_var = tk.BooleanVar(value=True)
        header_yes = tk.Radiobutton(
            frame_button, text="ヘッダ行あり", variable=self.header_var, value=True
        )
        header_yes.grid(row=0, column=3, padx=5, pady=5, sticky="w")

        header_no = tk.Radiobutton(
            frame_button, text="ヘッダ行なし", variable=self.header_var, value=False
        )
        header_no.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        header_yes.config(command=lambda: self.csv_import(False))
        header_no.config(command=lambda: self.csv_import(False))

        # 選択ラベルを作成
        file_label_title = tk.Label(frame_info, text="選択:", width=4, anchor="w")
        file_label_title.grid(row=0, column=0, padx=5, pady=1, sticky="w")
        self.file_label = tk.Label(frame_info, text="なし", width=80, anchor="w")
        self.file_label.grid(row=0, column=1, padx=0, pady=1, sticky="w")

        # 情報ラベルを作成
        info_label_title = tk.Label(frame_info, text="情報:", width=4, anchor="w")
        info_label_title.grid(row=1, column=0, padx=5, pady=1, sticky="w")
        self.info_label = tk.Label(frame_info, text="", width=80, anchor="w", fg="red")
        self.info_label.grid(row=1, column=1, padx=0, pady=1, sticky="w")

        # 日付ラベルを作成
        date_label = tk.Label(frame_manip, text="日付列:", width=5, anchor="w")
        date_label.grid(row=0, column=0, padx=2, pady=5, sticky="w")
        date_var = tk.StringVar(value="1")
        self.date_select = ttk.Combobox(
            frame_manip,
            textvariable=date_var,
            values=[str(i) for i in range(1, 10)],
            width=3,
        )
        self.date_select.grid(row=0, column=1, padx=2, pady=5, sticky="w")

        # 名目ラベルを作成
        comment_label = tk.Label(frame_manip, text="名目列:", width=5, anchor="w")
        comment_label.grid(row=0, column=2, padx=2, pady=5, sticky="w")
        comment_var = tk.StringVar(value="2")
        self.comment_select = ttk.Combobox(
            frame_manip,
            textvariable=comment_var,
            values=[str(i) for i in range(1, 10)],
            width=3,
        )
        self.comment_select.grid(row=0, column=3, padx=2, sticky="w")

        # 金額ラベルを作成
        pay_label = tk.Label(frame_manip, text="金額列:", width=5, anchor="w")
        pay_label.grid(row=0, column=4, padx=2, pady=5, sticky="w")
        pay_var = tk.StringVar(value="3")
        self.pay_select = ttk.Combobox(
            frame_manip,
            textvariable=pay_var,
            values=[str(i) for i in range(1, 10)],
            width=3,
        )
        self.pay_select.grid(row=0, column=5, padx=2, pady=5, sticky="w")

        # テンプレート選択ボックスを作成
        template_label = tk.Label(
            frame_manip, text="フォーマット:", width=7, anchor="w"
        )
        template_label.grid(row=0, column=6, padx=2, pady=5, sticky="w")
        self.template_var = tk.StringVar(value="三井住友")
        template_select = ttk.Combobox(
            frame_manip,
            textvariable=self.template_var,
            values=["三井住友", "amazon"],
            width=15,
        )
        template_select.grid(row=0, column=7, padx=2, pady=5, sticky="w")
        template_select.bind("<<ComboboxSelected>>", self.on_template_change)

        # Treeviewの設定
        frame_tree.columnconfigure(0, weight=1)  # 1列目を可変サイズとする
        frame_tree.rowconfigure(0, weight=1)  # 1行目を可変サイズとする
        frame_tree.grid_propagate(False)  # 内部のサイズに合わせたフレームサイズとしない

        # グリッドビューを初期化
        self.tree = ttk.Treeview(
            frame_tree, columns=("", "", ""), show="headings", height=15
        )
        self.tree.grid(row=0, column=0, columnspan=2, padx=5, pady=5)

        # スクロールバーを設定
        v_scrollbar = ttk.Scrollbar(
            frame_tree, orient=tk.VERTICAL, command=self.tree.yview
        )
        h_scrollbar = ttk.Scrollbar(
            frame_tree, orient=tk.HORIZONTAL, command=self.tree.xview
        )
        h_scrollbar.grid(row=1, column=0, sticky=tk.EW)
        v_scrollbar.grid(row=0, column=1, sticky=tk.NS)
        self.tree.configure(
            yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set
        )

        # メインループを開始
        root.mainloop()


if __name__ == "__main__":
    converter = CsvToExcelConverter()
    converter.init()
